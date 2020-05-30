def main() :
    #Getting input data
    D_raw,D_raw_1,D_raw_2,D_raw_3,D_raw_4 = Get_inputs()    
    #------------------------------------
    #General inoformation of datasets
    pure_users_accessed_index,D0 = General_info (D_raw) 
    #------------------------------------        
    #Characteristics
    D1, D2, D3, D4 = Charac_info (pure_users_accessed_index,D_raw,D_raw_1,D_raw_2,D_raw_3,D_raw_4)
    #------------------------------------
    #Cross Sectional Analysis
##    Cross_Sectional_Analysis (D1, D2, D3, D4)
    #------------------------------------
    #Gender, Country, and Sport in a Cross Sectional Analysis 
    G_Cross_Sectional_Analysis (D0, D1, D2, D3, D4)
    #------------------------------------
##    CS_Cross_Sectional_Analysis (D0 , D1, D2, D3, D4)
    #------------------------------------
##    Age_Cross_Sectional_Analysis (D0 , D1, D2, D3, D4)
    #------------------------------------
##    longitudinal_analysis_general (D0 , D1, D2, D3, D4)
    #------------------------------------
##    longitudinal_analysis_correlation (D0, D1, D2, D3, D4)

#--------------------
def Get_inputs() :
    
    import os #for giving adress to where we want to save output files
    from openpyxl import load_workbook
    import numpy as np
    
    wb_1 = load_workbook('OlympicAthletesData_1.xlsx')
    wb_2 = load_workbook('OlympicAthletesData_2.xlsx')
    wb_3 = load_workbook('OlympicAthletesData_3.xlsx')
    wb_4 = load_workbook('OlympicAthletesData_4.xlsx')
    
    ws_1 = wb_1.active
    ws_2 = wb_2.active
    ws_3 = wb_3.active
    ws_4 = wb_4.active
    
    Sport=[]
    Country=[]
    Name=[]
    Gender=[]
    ID=[]
    Age=[]
    
    N_Post_1=[]
    N_Follower_1=[]
    N_Following_1=[]
    Max_Like_1=[]
    Max_Comment_1=[]
    Self_Presenting_1=[]
    Pure_Self_Presenting_1=[]

    N_Post_2=[]
    N_Follower_2=[]
    N_Following_2=[]
    Max_Like_2=[]
    Max_Comment_2=[]
    Self_Presenting_2=[]
    Pure_Self_Presenting_2=[]



    N_Post_3=[]
    N_Follower_3=[]
    N_Following_3=[]
    Max_Like_3=[]
    Max_Comment_3=[]
    Self_Presenting_3=[]
    Pure_Self_Presenting_3=[]


    N_Post_4=[]
    N_Follower_4=[]
    N_Following_4=[]
    Max_Like_4=[]
    Max_Comment_4=[]
    Self_Presenting_4=[]
    Pure_Self_Presenting_4=[]

    #getting data of each column of excel file into a list named by the name of corrospondant name of that column 
    for row in ws_1 :
        N_Post_1.append(row[5].value)
        N_Follower_1.append(row[6].value)
        N_Following_1.append(row[7].value)
        Max_Like_1.append(row[8].value)
        Max_Comment_1.append(row[9].value)
        Self_Presenting_1.append(row[10].value)
        Pure_Self_Presenting_1.append(row[11].value)
        

    for row in ws_2 :
        N_Post_2.append(row[5].value)
        N_Follower_2.append(row[6].value)
        N_Following_2.append(row[7].value)
        Max_Like_2.append(row[8].value)
        Max_Comment_2.append(row[9].value)
        Self_Presenting_2.append(row[10].value)
        Pure_Self_Presenting_2.append(row[11].value)

    for row in ws_3 :
        N_Post_3.append(row[5].value)
        N_Follower_3.append(row[6].value)
        N_Following_3.append(row[7].value)
        Max_Like_3.append(row[8].value)
        Max_Comment_3.append(row[9].value)
        Self_Presenting_3.append(row[10].value)
        Pure_Self_Presenting_3.append(row[11].value)

    for row in ws_4 :
        Sport.append(row[0].value)
        Country.append(row[1].value)
        Name.append(row[2].value)
        Gender.append(row[3].value)
        ID.append(row[4].value)
        N_Post_4.append(row[5].value)
        N_Follower_4.append(row[6].value)
        N_Following_4.append(row[7].value)
        Max_Like_4.append(row[8].value)
        Max_Comment_4.append(row[9].value)
        Self_Presenting_4.append(row[10].value)
        Pure_Self_Presenting_4.append(row[11].value)
        Age.append(row[12].value)

    del Sport [0]
    del Country [0]
    del Name [0]
    del Gender [0]
    del ID [0]
    del Age [0]
    
    del N_Post_1 [0]
    del N_Follower_1 [0]
    del N_Following_1 [0]
    del Max_Like_1 [0]
    del Max_Comment_1 [0]
    del Self_Presenting_1 [0]
    del Pure_Self_Presenting_1 [0]

    del N_Post_2 [0]
    del N_Follower_2 [0]
    del N_Following_2 [0]
    del Max_Like_2 [0]
    del Max_Comment_2 [0]
    del Self_Presenting_2 [0]
    del Pure_Self_Presenting_2 [0]

    del N_Post_3 [0]
    del N_Follower_3 [0]
    del N_Following_3 [0]
    del Max_Like_3 [0]
    del Max_Comment_3 [0]
    del Self_Presenting_3 [0]
    del Pure_Self_Presenting_3 [0]

    del N_Post_4 [0]
    del N_Follower_4 [0]
    del N_Following_4 [0]
    del Max_Like_4 [0]
    del Max_Comment_4 [0]
    del Self_Presenting_4 [0]
    del Pure_Self_Presenting_4 [0]

    D_raw = np.array([Sport,Country,Name,Gender,ID,Age])
    D1_raw = np.array([N_Post_1,N_Follower_1,N_Following_1,Max_Like_1,Max_Comment_1,Self_Presenting_1,Pure_Self_Presenting_1])
    D2_raw = np.array([N_Post_2,N_Follower_2,N_Following_2,Max_Like_2,Max_Comment_2,Self_Presenting_2,Pure_Self_Presenting_2])
    D3_raw = np.array([N_Post_3,N_Follower_3,N_Following_3,Max_Like_3,Max_Comment_3,Self_Presenting_3,Pure_Self_Presenting_3])
    D4_raw = np.array([N_Post_4,N_Follower_4,N_Following_4,Max_Like_4,Max_Comment_4,Self_Presenting_4,Pure_Self_Presenting_4])



    return D_raw,D1_raw,D2_raw,D3_raw,D4_raw 
#-----------
def General_info (D_raw) :
    import numpy as np
    Sport_final=[]
    Country_final=[]
    Name_final=[]
    Gender_final=[]
    ID_final=[]
    Age_final=[]
    number_of_users_in_dataset = len (D_raw[2])
    pure_users = [] #users without their duplications (for example if a user has won more than one medal then the repititions would be omitted)
    pure_users_index = [] #pure users' index 
    pure_users_accessed_index = [] #pure users' index with their instagram account available 
    for n in D_raw[2]:
        if n not in pure_users :
            pure_users.append (n)
            z = list(D_raw[2]).index(n)
            pure_users_index.append(z)
            if D_raw[4][z]:
                pure_users_accessed_index.append(z)

    f_number = 0
    m_number = 0
    for i in pure_users_accessed_index :
        Sport_final.append(D_raw[0][i])
        Country_final.append(D_raw[1][i])
        Name_final.append(D_raw[2][i])
        Gender_final.append(D_raw[3][i])
        ID_final.append(D_raw[4][i])
        Age_final.append(D_raw[5][i])
        
        if D_raw[3][i] == "f" :
            f_number += 1
        else:
            m_number += 1

    D0_T = np.array ([Sport_final,Country_final,Name_final,Gender_final,ID_final,Age_final])
    D0 = D0_T.T
            
    print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print ("number_of_users_in_dataset (number of individual gold medals) : ", number_of_users_in_dataset)
    print ("number_of_pure_users (number of champions who won the gold medals (one champion might have won more than one gold medal) : ", len(pure_users))
    print ("number_of_pure_users_with_insta_account_available  : ", len(pure_users_accessed_index))
    print ("number of females : ", f_number)
    print ("number of males : ", m_number)
    print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    
    return pure_users_accessed_index,D0
#-----------
def Charac_info (pure_users_accessed_index,D_raw,D_raw_1,D_raw_2,D_raw_3,D_raw_4) :

    import numpy as np
    
    ratio_of_followings_to_followers_1  = []
    ratio_of_self_presenting_1 = []
    ratio_of_pure_self_presenting_to_self_presenting_1 = []
    ratio_of_engagement_to_followers_1 = []
    Age_1 = []

    ratio_of_followings_to_followers_2  = []
    ratio_of_self_presenting_2 = []
    ratio_of_pure_self_presenting_to_self_presenting_2 = []
    ratio_of_engagement_to_followers_2 = []
    Age_2 = []

    ratio_of_followings_to_followers_3  = []
    ratio_of_self_presenting_3 = []
    ratio_of_pure_self_presenting_to_self_presenting_3 = []
    ratio_of_engagement_to_followers_3 = []
    Age_3 = []

    ratio_of_followings_to_followers_4  = []
    ratio_of_self_presenting_4 = []
    ratio_of_pure_self_presenting_to_self_presenting_4 = []
    ratio_of_engagement_to_followers_4 = []
    Age_4 = []
    
    D_number = (1,2,3,4)
    for dn in D_number:
        for i in pure_users_accessed_index:
            eval ('ratio_of_followings_to_followers_' + str(dn)).append(eval('D_raw_' + str(dn))[2][i]/eval('D_raw_' + str(dn))[1][i])
            eval ('ratio_of_self_presenting_' + str(dn)).append(eval('D_raw_' + str(dn))[5][i]/10)
            eval ('ratio_of_pure_self_presenting_to_self_presenting_' + str(dn)).append(eval('D_raw_' + str(dn))[6][i]/eval('D_raw_' + str(dn))[5][i])
            engagement = eval('D_raw_'+str(dn))[3][i] + eval('D_raw_'+str(dn))[4][i]
            eval ('ratio_of_engagement_to_followers_' + str(dn)).append(engagement/eval('D_raw_' + str(dn))[1][i])
            eval ('Age_' + str(dn)).append(D_raw[5][i])
            
        D1_T = np.array ([ratio_of_followings_to_followers_1,ratio_of_self_presenting_1,ratio_of_pure_self_presenting_to_self_presenting_1,ratio_of_engagement_to_followers_1,Age_1])
        D2_T = np.array ([ratio_of_followings_to_followers_2,ratio_of_self_presenting_2,ratio_of_pure_self_presenting_to_self_presenting_2,ratio_of_engagement_to_followers_2,Age_2])
        D3_T = np.array ([ratio_of_followings_to_followers_3,ratio_of_self_presenting_3,ratio_of_pure_self_presenting_to_self_presenting_3,ratio_of_engagement_to_followers_3,Age_3])
        D4_T = np.array ([ratio_of_followings_to_followers_4,ratio_of_self_presenting_4,ratio_of_pure_self_presenting_to_self_presenting_4,ratio_of_engagement_to_followers_4,Age_4])

        #now we transpose the matrices
        D1 = D1_T.T
        D2 = D2_T.T
        D3 = D3_T.T
        D4 = D4_T.T
        
    #here we have 4 matrices (D1,D2,D3,D4) of 144*5, each one achieved from an observation in a specific time (with one month distance between each observation)
    # 5 is the number of attributes and 144 is the number of olympic champions, so each matrix is as follow:
    
    #                                  fw/fr       self-presenting (sp)      psp/sp    engagement/follower (en/fr)      age
    #user1                               x1               y1                    z1              h1                       f1
    #user2                               x2               y1                    z1              h1                       f1              
    #.                                   .                .                     .               .                         .     
    #.                                   .                .                     .               .                         .     
    #user144                            x144            y144                    z144            h144                    f144     

    
    return D1, D2, D3, D4
#-----------
def Cross_Sectional_Analysis (D1, D2, D3, D4):
    import numpy as np
    import pandas as pd
    from scipy.stats import spearmanr
    from scipy.stats import pearsonr
    import seaborn as sns
    import matplotlib.pyplot as plt


    df1 = pd.DataFrame(D1, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age'])
    df2 = pd.DataFrame(D2, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age'])
    df3 = pd.DataFrame(D3, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age'])
    df4 = pd.DataFrame(D4, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age'])

    

    #D1------------------------
    coeffmat = np.zeros((df1.shape[1], df1.shape[1]))
    pvalmat = np.zeros((df1.shape[1], df1.shape[1]))
    for i in range(df1.shape[1]):    
        for j in range(df1.shape[1]):        
            corrtest = pearsonr(df1[df1.columns[i]], df1[df1.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df1.columns, index=df1.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df1.columns, index=df1.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.triu_indices_from(mask, k=1)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D1_pearson.png', dpi=500)    

    coeffmat = np.zeros((df1.shape[1], df1.shape[1]))
    pvalmat = np.zeros((df1.shape[1], df1.shape[1]))
    for i in range(df1.shape[1]):    
        for j in range(df1.shape[1]):        
            corrtest = spearmanr(df1[df1.columns[i]], df1[df1.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df1.columns, index=df1.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df1.columns, index=df1.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.tril_indices_from(mask)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D1_Spearman.png', dpi=500)


    
    #D2------------------------
    coeffmat = np.zeros((df2.shape[1], df2.shape[1]))
    pvalmat = np.zeros((df2.shape[1], df2.shape[1]))
    for i in range(df2.shape[1]):    
        for j in range(df2.shape[1]):        
            corrtest = pearsonr(df2[df2.columns[i]], df2[df2.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df2.columns, index=df2.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df2.columns, index=df2.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.triu_indices_from(mask, k=1)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D2_pearson.png', dpi=500)    

    coeffmat = np.zeros((df2.shape[1], df2.shape[1]))
    pvalmat = np.zeros((df2.shape[1], df2.shape[1]))
    for i in range(df2.shape[1]):    
        for j in range(df2.shape[1]):        
            corrtest = spearmanr(df2[df2.columns[i]], df2[df2.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df2.columns, index=df2.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df2.columns, index=df2.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.tril_indices_from(mask)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D2_Spearman.png', dpi=500)


    
    #D3------------------------
    coeffmat = np.zeros((df3.shape[1], df3.shape[1]))
    pvalmat = np.zeros((df3.shape[1], df3.shape[1]))
    for i in range(df3.shape[1]):    
        for j in range(df3.shape[1]):        
            corrtest = pearsonr(df3[df3.columns[i]], df3[df3.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df3.columns, index=df3.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df3.columns, index=df3.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.triu_indices_from(mask, k=1)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D3_pearson.png', dpi=500)    

    coeffmat = np.zeros((df3.shape[1], df3.shape[1]))
    pvalmat = np.zeros((df3.shape[1], df3.shape[1]))
    for i in range(df3.shape[1]):    
        for j in range(df3.shape[1]):        
            corrtest = spearmanr(df3[df3.columns[i]], df3[df3.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df3.columns, index=df3.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df3.columns, index=df3.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.tril_indices_from(mask)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D3_Spearman.png', dpi=500)

    
    #D4------------------------
    coeffmat = np.zeros((df4.shape[1], df4.shape[1]))
    pvalmat = np.zeros((df4.shape[1], df4.shape[1]))
    for i in range(df4.shape[1]):    
        for j in range(df4.shape[1]):        
            corrtest = pearsonr(df4[df4.columns[i]], df4[df4.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df4.columns, index=df4.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df4.columns, index=df4.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.triu_indices_from(mask, k=1)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D4_pearson.png', dpi=500)    

    coeffmat = np.zeros((df4.shape[1], df4.shape[1]))
    pvalmat = np.zeros((df4.shape[1], df4.shape[1]))
    for i in range(df4.shape[1]):    
        for j in range(df4.shape[1]):        
            corrtest = spearmanr(df4[df4.columns[i]], df4[df4.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df4.columns, index=df4.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df4.columns, index=df4.columns)
    cmap = sns.diverging_palette(240, 10, s=80, l=55, as_cmap=True)
    f, ax = plt.subplots(figsize=(9, 6))
    mask = np.zeros_like(dfcoeff)
    mask[np.tril_indices_from(mask)] = True
    ax = sns.heatmap(dfpvals, mask=mask, vmin=0, vmax=0.09, cmap=cmap, center=0.05, annot=dfcoeff, annot_kws={'size':16}, square=True, linewidths=.5, ax = ax)
    plt.yticks(fontsize=16, rotation=0)
    plt.xticks(fontsize=16)
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=16)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\D4_Spearman.png', dpi=500)

#-----------
#cross sectional analysis over gender
def G_Cross_Sectional_Analysis (D0, D1, D2, D3, D4):
    import seaborn as sns
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd

    index_m = np.where(D0[:,3] == 'm')
    index_f = np.where(D0[:,3] == 'f')
    for i in index_m:
        fw_fr_1_m = D1[:,0][i]
        sp_1_m = D1[:,1][i]
        psp_sp_1_m = D1[:,2][i]
        en_fr_1_m = D1[:,3][i]
        fw_fr_2_m = D2[:,0][i]
        sp_2_m = D2[:,1][i]
        psp_sp_2_m = D2[:,2][i]
        en_fr_2_m = D2[:,3][i]
        fw_fr_3_m = D3[:,0][i]
        sp_3_m = D3[:,1][i]
        psp_sp_3_m = D3[:,2][i]
        en_fr_3_m = D3[:,3][i]
        fw_fr_4_m = D4[:,0][i]
        sp_4_m = D4[:,1][i]
        psp_sp_4_m = D4[:,2][i]
        en_fr_4_m = D4[:,3][i]
        age_m = D4[:,4][i]

    for i in index_f:
        fw_fr_1_f = D1[:,0][i]
        sp_1_f = D1[:,1][i]
        psp_sp_1_f = D1[:,2][i]
        en_fr_1_f = D1[:,3][i]
        fw_fr_2_f = D2[:,0][i]
        sp_2_f = D2[:,1][i]
        psp_sp_2_f = D2[:,2][i]
        en_fr_2_f = D2[:,3][i]
        fw_fr_3_f = D3[:,0][i]
        sp_3_f = D3[:,1][i]
        psp_sp_3_f = D3[:,2][i]
        en_fr_3_f = D3[:,3][i]
        fw_fr_4_f = D4[:,0][i]
        sp_4_f = D4[:,1][i]
        psp_sp_4_f = D4[:,2][i]
        en_fr_4_f = D4[:,3][i]
        age_f = D4[:,4][i]
    
    fw_fr_all_m = np.concatenate((fw_fr_1_m, fw_fr_2_m, fw_fr_3_m, fw_fr_4_m))
    sp_all_m = np.concatenate((sp_1_m, sp_2_m, sp_3_m, sp_4_m))
    psp_sp_all_m = np.concatenate((psp_sp_1_m, psp_sp_2_m, psp_sp_3_m, psp_sp_4_m))
    en_fr_all_m = np.concatenate((en_fr_1_m, en_fr_2_m, en_fr_3_m, en_fr_4_m))
    age_all_m = np.concatenate((age_m, age_m, age_m, age_m))

    fw_fr_all_f = np.concatenate((fw_fr_1_f, fw_fr_2_f, fw_fr_3_f, fw_fr_4_f))
    sp_all_f = np.concatenate((sp_1_f, sp_2_f, sp_3_f, sp_4_f))
    psp_sp_all_f = np.concatenate((psp_sp_1_f, psp_sp_2_f, psp_sp_3_f, psp_sp_4_f))
    en_fr_all_f = np.concatenate((en_fr_1_f, en_fr_2_f, en_fr_3_f, en_fr_4_f))
    age_all_f = np.concatenate((age_f, age_f, age_f, age_f))

    sns.set(style="darkgrid")

    #---- en/fr & sp
    
    #--male
    f, ax = plt.subplots(figsize=(12, 10))
    data = pd.DataFrame(np.array ([en_fr_all_m,sp_all_m]).T,columns=['x', 'y'])
    ax = sns.jointplot('x', 'y' ,data ,kind="kde", space=0, color="b", n_levels=60)
    ax.ax_joint.collections[0].set_alpha(0)
    ax.set_axis_labels("en/fr", "sp", fontsize=24)
    rc={'xtick.labelsize': 18, 'ytick.labelsize': 18}
    plt.rcParams.update(**rc)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Gender_CS\\en-fr--sp_m.png', dpi=500)

    #--female
    f, ax = plt.subplots(figsize=(12, 10))
    data = pd.DataFrame(np.array ([en_fr_all_f,sp_all_f]).T,columns=['x', 'y'])
    ax = sns.jointplot('x', 'y', data, kind="kde", space=0, color="r", n_levels=60)
    ax.ax_joint.collections[0].set_alpha(0)
    ax.set_axis_labels("en/fr", "sp", fontsize=24)
    rc={'xtick.labelsize': 18, 'ytick.labelsize': 18}
    plt.rcParams.update(**rc)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Gender_CS\\en-fr--sp_f.png', dpi=500)
    

    #---- en/fr & fw/fr
    #--male
    f, ax = plt.subplots(figsize=(12, 10))
    data = pd.DataFrame(np.array ([en_fr_all_m,fw_fr_all_m]).T,columns=['x', 'y'])
    ax = sns.jointplot('x', 'y' ,data ,kind="kde", space=0, color="b", n_levels=60)
    ax.ax_joint.collections[0].set_alpha(0)
    ax.set_axis_labels("en/fr", "fw/fr", fontsize=24)
    rc={'xtick.labelsize': 18, 'ytick.labelsize': 18}
    plt.rcParams.update(**rc)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Gender_CS\\en-fr--fw-fr_m.png', dpi=500)
    
    #--female
    f, ax = plt.subplots(figsize=(12, 10))
    data = pd.DataFrame(np.array ([en_fr_all_f,fw_fr_all_f]).T,columns=['x', 'y'])
    ax = sns.jointplot('x', 'y', data, kind="kde", space=0, color="r", n_levels=60)
    ax.ax_joint.collections[0].set_alpha(0)
    ax.set_axis_labels("en/fr", "fw/fr", fontsize=24)
    rc={'xtick.labelsize': 18, 'ytick.labelsize': 18}
    plt.rcParams.update(**rc)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Gender_CS\\en-fr--fw-fr_f.png', dpi=500)


    #---- psp/sp & age
    #--male
    f, ax = plt.subplots(figsize=(12, 10))
    data = pd.DataFrame(np.array ([psp_sp_all_m,age_all_m]).T,columns=['x', 'y'])
    ax = sns.jointplot('x', 'y', data, kind="kde", space=0, color="b", n_levels=60)
    ax.ax_joint.collections[0].set_alpha(0)
    ax.set_axis_labels("psp/sp", "age", fontsize=24)
    rc={'xtick.labelsize': 18, 'ytick.labelsize': 18}
    plt.rcParams.update(**rc)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Gender_CS\\psp-sp--age_m.png', dpi=500)

    #--female
    f, ax = plt.subplots(figsize=(12, 10))
    data = pd.DataFrame(np.array ([psp_sp_all_f,age_all_f]).T,columns=['x', 'y'])
    ax = sns.jointplot('x', 'y', data, kind="kde", space=0, color="r", n_levels=60)
    ax.ax_joint.collections[0].set_alpha(0)
    ax.set_axis_labels("psp/sp", "age", fontsize=24)
    rc={'xtick.labelsize': 18, 'ytick.labelsize': 18}
    plt.rcParams.update(**rc)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Gender_CS\\psp-sp--age_f.png', dpi=500)

#-----------
#cross sectional analysis over country and sports
def CS_Cross_Sectional_Analysis (D0 , D1, D2, D3, D4):
    import seaborn as sns
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd

    #Countries 
    Countries = set(D0[:,1])
    D_all = np.concatenate((D1, D2, D3, D4))
    D0_all = np.concatenate((D0, D0, D0, D0))

    fw_fr_all_countries = []
    sp_all_countries = []
    psp_sp_all_countries = []
    en_fr_all_countries = []

    index_c_all_countries = []
    
    for c in Countries:
        index_c = np.where(D0_all[:,1] == c)
        index_c_all_countries.append (index_c)
        for i in index_c:
            fw_fr = D_all[:,0][i]            
            sp = D_all[:,1][i]
            psp_sp = D_all[:,2][i]
            en_fr = D_all[:,3][i]
            
            fw_fr_all_countries.append (fw_fr)
            sp_all_countries.append (sp)
            psp_sp_all_countries.append (psp_sp)
            en_fr_all_countries.append (en_fr)

                
    fw_fr_mean = pd.DataFrame(fw_fr_all_countries).mean(axis = 1)
    fw_fr_var = pd.DataFrame(fw_fr_all_countries).var(axis = 1)
    sp_mean = pd.DataFrame(sp_all_countries).mean(axis = 1)
    sp_var = pd.DataFrame(sp_all_countries).var(axis = 1)
    psp_sp_mean = pd.DataFrame(psp_sp_all_countries).mean(axis = 1)
    psp_sp_var = pd.DataFrame(psp_sp_all_countries).var(axis = 1)
    en_fr_mean = pd.DataFrame(en_fr_all_countries).mean(axis = 1)
    en_fr_var = pd.DataFrame(en_fr_all_countries).var(axis = 1)


    #Countries (the 5 toppest and lowest countries in each attribute)
    f = open('Country&Sport.txt','a')
    f.write('----------------')
    f.write('----------------')
    f.write('----------------')
    f.write('C O U N T R Y')
    f.write('----------------')
    f.write('----------------')
    f.write('----------------')
    f.write('\n' + '     '  +  'fw/fr')
    f.write('\n' + '----------------')
    f.write('\n' + 'Country   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in fw_fr_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(fw_fr_mean[i])  +  '    ' + str(fw_fr_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in fw_fr_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(fw_fr_mean[i])  +  '    ' + str(fw_fr_var[i]))
    f.write('\n' + '----------------')            

    f.write('\n' + '----------------')
    f.write('\n' + '     '  +  'sp')
    f.write('\n' + '----------------')
    f.write('\n' + 'Country   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in sp_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(sp_mean[i])  +  '    ' + str(sp_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in sp_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(sp_mean[i])  +  '    ' + str(sp_var[i]))
    f.write('\n' + '----------------')

    f.write('\n' + '----------------')
    f.write('\n' + '     '  +  'psp/sp')
    f.write('\n' + '----------------')
    f.write('\n' + 'Country   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in psp_sp_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(psp_sp_mean[i])  +  '    ' + str(psp_sp_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in psp_sp_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(psp_sp_mean[i])  +  '    ' + str(psp_sp_var[i]))
    f.write('\n' + '----------------')

    f.write('\n' + '----------------')
    f.write('\n' + '     '  +  'en/fr')
    f.write('\n' + '----------------')
    f.write('\n' + 'Country   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in en_fr_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(en_fr_mean[i])  +  '    ' + str(en_fr_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in en_fr_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Countries)[i])   +  '    ' +  str(en_fr_mean[i])  +  '    ' + str(en_fr_var[i]))
    f.write('\n' + '----------------')

    #Sports 
    Sports = set(D0[:,0])
    D_all = np.concatenate((D1, D2, D3, D4))
    D0_all = np.concatenate((D0, D0, D0, D0))

    fw_fr_all_sports = []
    sp_all_sports = []
    psp_sp_all_sports = []
    en_fr_all_sports = []

    index_s_all_sports = []
    
    for s in Sports:
        index_s = np.where(D0_all[:,0] == s)
        index_s_all_sports.append (index_s)
        for i in index_s:
            fw_fr = D_all[:,0][i]            
            sp = D_all[:,1][i]
            psp_sp = D_all[:,2][i]
            en_fr = D_all[:,3][i]
            
            fw_fr_all_sports.append (fw_fr)
            sp_all_sports.append (sp)
            psp_sp_all_sports.append (psp_sp)
            en_fr_all_sports.append (en_fr)

                
    fw_fr_mean = pd.DataFrame(fw_fr_all_sports).mean(axis = 1)
    fw_fr_var = pd.DataFrame(fw_fr_all_sports).var(axis = 1)
    sp_mean = pd.DataFrame(sp_all_sports).mean(axis = 1)
    sp_var = pd.DataFrame(sp_all_sports).var(axis = 1)
    psp_sp_mean = pd.DataFrame(psp_sp_all_sports).mean(axis = 1)
    psp_sp_var = pd.DataFrame(psp_sp_all_sports).var(axis = 1)
    en_fr_mean = pd.DataFrame(en_fr_all_sports).mean(axis = 1)
    en_fr_var = pd.DataFrame(en_fr_all_sports).var(axis = 1)


    #Sports (the 5 toppest and lowest countries in each attribute)
    f.write('----------------')
    f.write('----------------')
    f.write('----------------')
    f.write('S P O R T S')
    f.write('----------------')
    f.write('----------------')
    f.write('----------------')
    f.write('\n' + '     '  +  'fw/fr')
    f.write('\n' + '----------------')
    f.write('\n' + 'Sport   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in fw_fr_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(fw_fr_mean[i])  +  '    ' + str(fw_fr_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in fw_fr_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(fw_fr_mean[i])  +  '    ' + str(fw_fr_var[i]))
    f.write('\n' + '----------------')            

    f.write('\n' + '----------------')
    f.write('\n' + '     '  +  'sp')
    f.write('\n' + '----------------')
    f.write('\n' + 'Sport   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in sp_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(sp_mean[i])  +  '    ' + str(sp_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in sp_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(sp_mean[i])  +  '    ' + str(sp_var[i]))
    f.write('\n' + '----------------')

    f.write('\n' + '----------------')
    f.write('\n' + '     '  +  'psp/sp')
    f.write('\n' + '----------------')
    f.write('\n' + 'Sport   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in psp_sp_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(psp_sp_mean[i])  +  '    ' + str(psp_sp_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in psp_sp_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(psp_sp_mean[i])  +  '    ' + str(psp_sp_var[i]))
    f.write('\n' + '----------------')

    f.write('\n' + '----------------')
    f.write('\n' + '     '  +  'en/fr')
    f.write('\n' + '----------------')
    f.write('\n' + 'Sport   ,  mean   ,   var')
    f.write('\n' + '---- L A R G E S T ------')
    for i in en_fr_mean.nlargest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(en_fr_mean[i])  +  '    ' + str(en_fr_var[i]))
    f.write('\n' + '---- S M A L L E S T-----')
    for i in en_fr_mean.nsmallest(5).index.values:
        f.write('\n' + str(list(Sports)[i])   +  '    ' +  str(en_fr_mean[i])  +  '    ' + str(en_fr_var[i]))
    f.write('\n' + '----------------')

    f.close()

#-----------
#cross sectional analysis over AGE and Gender
def Age_Cross_Sectional_Analysis (D0 , D1, D2, D3, D4):
    import seaborn as sns
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd


    D_all = np.concatenate((D1, D2, D3, D4))
    D0_all = np.concatenate((D0, D0, D0, D0))
    
    #-------------
    #Age Seperation
    ages = D0_all[:,5].astype(int)            
        
    new_column = np.array ([])
    for a in ages:
        if a < 26:
            new_column = np.append (new_column, 'R1')
        elif a < 30:
            new_column = np.append (new_column, 'R2')
        else:
            new_column = np.append (new_column, 'R3')
        
    D_all = pd.DataFrame(D_all, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #576*5
    D_all.insert(5, 'Age_Range', new_column)
    #-------------
    #Gender Seperation
    new_column_g = np.array ([])
    for g in D0_all[:,3]:
        if g == 'm':
            new_column_g = np.append (new_column_g, 'Male')
        else:
            new_column_g = np.append (new_column_g, 'Female')

    D_all.insert(6, 'Gender', new_column_g)
    #-------------

    sns.set(style="whitegrid")
    
    fig, ax = plt.subplots()
    ax = sns.violinplot(x='Age_Range', y='fw/fr', data=D_all, order=["R1", "R2", "R3"], hue="Gender", split=True, palette="Set2")
    plt.xlabel('Age_Range', fontsize=22)
    plt.ylabel('fw/fr', fontsize=22)
    plt.yticks(fontsize=20)
    plt.xticks(fontsize=20)
    plt.gcf().subplots_adjust(left=0.18)
    plt.gcf().subplots_adjust(bottom=0.18)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Age_Gender\\Age_fw-fr.png', dpi=500)
    
    
    fig, ax = plt.subplots()
    ax = sns.violinplot(x='Age_Range', y='sp', data=D_all, order=["R1", "R2", "R3"], hue="Gender", split=True, palette="Set2")
    plt.xlabel('Age_Range', fontsize=22)
    plt.ylabel('sp', fontsize=22)
    plt.yticks(fontsize=20)
    plt.xticks(fontsize=20)
    plt.gcf().subplots_adjust(left=0.18)
    plt.gcf().subplots_adjust(bottom=0.18)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Age_Gender\\Age_sp.png', dpi=500)

    
    fig, ax = plt.subplots()
    ax = sns.violinplot(x='Age_Range', y='psp/sp', data=D_all, order=["R1", "R2", "R3"], hue="Gender", split=True, palette="Set2")
    plt.xlabel('Age_Range', fontsize=22)
    plt.ylabel('psp/sp', fontsize=22)
    plt.yticks(fontsize=20)
    plt.xticks(fontsize=20)
    plt.gcf().subplots_adjust(left=0.2)
    plt.gcf().subplots_adjust(bottom=0.18)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Age_Gender\\Age_psp-sp.png', dpi=500)

    
    fig, ax = plt.subplots()
    ax = sns.violinplot(x='Age_Range', y='en/fr', data=D_all, order=["R1", "R2", "R3"], hue="Gender", split=True, palette="Set2")
    plt.xlabel('Age_Range', fontsize=22)
    plt.ylabel('en/fr', fontsize=22)
    plt.yticks(fontsize=20)
    plt.xticks(fontsize=20)
    plt.gcf().subplots_adjust(left=0.18)
    plt.gcf().subplots_adjust(bottom=0.18)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Age_Gender\\Age_en-fr.png', dpi=500)

#-----------
def longitudinal_analysis__general (D0 , D1, D2, D3, D4):
    import seaborn as sns
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    
    df0 = pd.DataFrame(D0, columns=['Sport', 'Country', 'Name', 'Gender', 'ID', 'Age'])  #144*6
    df1 = pd.DataFrame(D1, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5
    df2 = pd.DataFrame(D2, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5
    df3 = pd.DataFrame(D3, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5
    df4 = pd.DataFrame(D4, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5

    #---------------------------------------------------
    #General Information about change in the characteristics
    #---------------------------------------------------
    index_m = np.where(D0[:,3] == 'm')
    index_f = np.where(D0[:,3] == 'f')
    for i in index_m:
        fw_fr_1_m = D1[:,0][i]
        sp_1_m = D1[:,1][i]
        psp_sp_1_m = D1[:,2][i]
        en_fr_1_m = D1[:,3][i]
        fw_fr_2_m = D2[:,0][i]
        sp_2_m = D2[:,1][i]
        psp_sp_2_m = D2[:,2][i]
        en_fr_2_m = D2[:,3][i]
        fw_fr_3_m = D3[:,0][i]
        sp_3_m = D3[:,1][i]
        psp_sp_3_m = D3[:,2][i]
        en_fr_3_m = D3[:,3][i]
        fw_fr_4_m = D4[:,0][i]
        sp_4_m = D4[:,1][i]
        psp_sp_4_m = D4[:,2][i]
        en_fr_4_m = D4[:,3][i]

    for i in index_f:
        fw_fr_1_f = D1[:,0][i]
        sp_1_f = D1[:,1][i]
        psp_sp_1_f = D1[:,2][i]
        en_fr_1_f = D1[:,3][i]
        fw_fr_2_f = D2[:,0][i]
        sp_2_f = D2[:,1][i]
        psp_sp_2_f = D2[:,2][i]
        en_fr_2_f = D2[:,3][i]
        fw_fr_3_f = D3[:,0][i]
        sp_3_f = D3[:,1][i]
        psp_sp_3_f = D3[:,2][i]
        en_fr_3_f = D3[:,3][i]
        fw_fr_4_f = D4[:,0][i]
        sp_4_f = D4[:,1][i]
        psp_sp_4_f = D4[:,2][i]
        en_fr_4_f = D4[:,3][i]


    
    #fw/fr
    N = 4
    allMeans = (df1['fw/fr'].mean(),df2['fw/fr'].mean(),df3['fw/fr'].mean(),df4['fw/fr'].mean())
    allStd = (df1['fw/fr'].var(),df2['fw/fr'].var(),df3['fw/fr'].var(),df4['fw/fr'].var())
    menMeans = (np.mean(fw_fr_1_m), np.mean(fw_fr_2_m), np.mean(fw_fr_3_m), np.mean(fw_fr_4_m))
    menStd = (np.var(fw_fr_1_m), np.var(fw_fr_2_m), np.var(fw_fr_3_m), np.var(fw_fr_4_m))
    womenMeans = (np.mean(fw_fr_1_f), np.mean(fw_fr_2_f), np.mean(fw_fr_3_f), np.mean(fw_fr_4_f))
    womenStd = ((np.var(fw_fr_1_f), np.var(fw_fr_2_f), np.var(fw_fr_3_f), np.var(fw_fr_4_f)))
    
    fig, ax = plt.subplots()
    ind = np.arange(N)
    width = 0.2
    p1 = ax.bar(ind, allMeans, width,  yerr=allStd)
    p2 = ax.bar(ind + width, menMeans, width,  yerr=menStd)
    p3 = ax.bar(ind + width + width, womenMeans, width, yerr=womenStd)
    ax.set_xticks(ind + width)
    ax.set_xticklabels(('D1', 'D2', 'D3', 'D4'))
    ax.legend((p1[0], p2[0], p3[0]), ('All', 'Men', 'Women'), loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.ylabel('fw/fr', fontsize=20)
    plt.yticks(fontsize=16)
    plt.xticks(fontsize=16)
    ax.autoscale_view()
    plt.gcf().subplots_adjust(left=0.15)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Longitudinal_Study\\fw-fr.png', dpi=500)


    #sp
    N = 4
    allMeans = (df1['sp'].mean(),df2['sp'].mean(),df3['sp'].mean(),df4['sp'].mean())
    allStd = (df1['sp'].var(),df2['sp'].var(),df3['sp'].var(),df4['sp'].var())
    menMeans = (np.mean(sp_1_m), np.mean(sp_2_m), np.mean(sp_3_m), np.mean(sp_4_m))
    menStd = (np.var(sp_1_m), np.var(sp_2_m), np.var(sp_3_m), np.var(sp_4_m))
    womenMeans = (np.mean(sp_1_f), np.mean(sp_2_f), np.mean(sp_3_f), np.mean(sp_4_f))
    womenStd = ((np.var(sp_1_f), np.var(sp_2_f), np.var(sp_3_f), np.var(sp_4_f)))
    
    fig, ax = plt.subplots()
    ind = np.arange(N)
    width = 0.2
    p1 = ax.bar(ind, allMeans, width,  yerr=allStd)
    p2 = ax.bar(ind + width, menMeans, width,  yerr=menStd)
    p3 = ax.bar(ind + width + width, womenMeans, width, yerr=womenStd)
    ax.set_xticks(ind + width)
    ax.set_xticklabels(('D1', 'D2', 'D3', 'D4'))
    ax.legend((p1[0], p2[0], p3[0]), ('All', 'Men', 'Women'), loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.ylabel('sp', fontsize=20)
    plt.yticks(fontsize=16)
    plt.xticks(fontsize=16)
    ax.autoscale_view()    
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Longitudinal_Study\\sp.png', dpi=500)


    #psp/sp
    N = 4
    allMeans = (df1['psp/sp'].mean(),df2['psp/sp'].mean(),df3['psp/sp'].mean(),df4['psp/sp'].mean())
    allStd = (df1['psp/sp'].var(),df2['psp/sp'].var(),df3['psp/sp'].var(),df4['psp/sp'].var())
    menMeans = (np.mean(psp_sp_1_m), np.mean(psp_sp_2_m), np.mean(psp_sp_3_m), np.mean(psp_sp_4_m))
    menStd = (np.var(psp_sp_1_m), np.var(psp_sp_2_m), np.var(psp_sp_3_m), np.var(psp_sp_4_m))
    womenMeans = (np.mean(psp_sp_1_f), np.mean(psp_sp_2_f), np.mean(psp_sp_3_f), np.mean(psp_sp_4_f))
    womenStd = ((np.var(psp_sp_1_f), np.var(psp_sp_2_f), np.var(psp_sp_3_f), np.var(psp_sp_4_f)))
    
    fig, ax = plt.subplots()
    ind = np.arange(N)
    width = 0.2
    p1 = ax.bar(ind, allMeans, width,  yerr=allStd)
    p2 = ax.bar(ind + width, menMeans, width,  yerr=menStd)
    p3 = ax.bar(ind + width + width, womenMeans, width, yerr=womenStd)
    ax.set_xticks(ind + width)
    ax.set_xticklabels(('D1', 'D2', 'D3', 'D4'))
    ax.legend((p1[0], p2[0], p3[0]), ('All', 'Men', 'Women') ,loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.ylabel('psp/sp', fontsize=20)
    plt.yticks(fontsize=16)
    plt.xticks(fontsize=16)
    ax.autoscale_view()
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Longitudinal_Study\\psp-sp.png', dpi=500)


    #en/fr
    N = 4
    allMeans = (df1['en/fr'].mean(),df2['en/fr'].mean(),df3['en/fr'].mean(),df4['en/fr'].mean())
    allStd = (df1['en/fr'].var(),df2['en/fr'].var(),df3['en/fr'].var(),df4['en/fr'].var())
    menMeans = (np.mean(en_fr_1_m), np.mean(en_fr_2_m), np.mean(en_fr_3_m), np.mean(en_fr_4_m))
    menStd = (np.var(en_fr_1_m), np.var(en_fr_2_m), np.var(en_fr_3_m), np.var(en_fr_4_m))
    womenMeans = (np.mean(en_fr_1_f), np.mean(en_fr_2_f), np.mean(en_fr_3_f), np.mean(en_fr_4_f))
    womenStd = ((np.var(en_fr_1_f), np.var(en_fr_2_f), np.var(en_fr_3_f), np.var(en_fr_4_f)))
    
    fig, ax = plt.subplots()
    ind = np.arange(N)
    width = 0.2
    p1 = ax.bar(ind, allMeans, width,  yerr=allStd)
    p2 = ax.bar(ind + width, menMeans, width,  yerr=menStd)
    p3 = ax.bar(ind + width + width, womenMeans, width, yerr=womenStd)
    ax.set_xticks(ind + width)
    ax.set_xticklabels(('D1', 'D2', 'D3', 'D4'))
    ax.legend((p1[0], p2[0], p3[0]), ('All', 'Men', 'Women'), loc='upper center', bbox_to_anchor=(0.5, 1.05), fancybox=True, shadow=True, ncol=5)
    plt.ylabel('en/fr', fontsize=20)
    plt.yticks(fontsize=16)
    plt.xticks(fontsize=16)
    ax.autoscale_view()
    plt.gcf().subplots_adjust(left=0.18)
    plt.savefig('D:\\Research Life\\Papers\\Paper_7\\Figures\\Longitudinal_Study\\en-fr.png', dpi=500)

#---------------------------------------------------
def longitudinal_analysis_correlation (D0, D1, D2, D3, D4):
    import numpy as np
    import pandas as pd

    df1 = pd.DataFrame(D1, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5
    df2 = pd.DataFrame(D2, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5
    df3 = pd.DataFrame(D3, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5
    df4 = pd.DataFrame(D4, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr', 'age']) #144*5


    df1 = df1.drop(['age'], axis = 1) #because this column (age) is consistent and equals to zero
    df2 = df2.drop(['age'], axis = 1) #because this column (age) is consistent and equals to zero
    df3 = df3.drop(['age'], axis = 1) #because this column (age) is consistent and equals to zero
    df4 = df4.drop(['age'], axis = 1) #because this column (age) is consistent and equals to zero

    #Gender Seperation
    index_m = np.where(D0[:,3] == 'm')
    index_f = np.where(D0[:,3] == 'f')
    for i in index_m:
        fw_fr_1_m = D1[:,0][i]
        sp_1_m = D1[:,1][i]
        psp_sp_1_m = D1[:,2][i]
        en_fr_1_m = D1[:,3][i]
        fw_fr_2_m = D2[:,0][i]
        sp_2_m = D2[:,1][i]
        psp_sp_2_m = D2[:,2][i]
        en_fr_2_m = D2[:,3][i]
        fw_fr_3_m = D3[:,0][i]
        sp_3_m = D3[:,1][i]
        psp_sp_3_m = D3[:,2][i]
        en_fr_3_m = D3[:,3][i]
        fw_fr_4_m = D4[:,0][i]
        sp_4_m = D4[:,1][i]
        psp_sp_4_m = D4[:,2][i]
        en_fr_4_m = D4[:,3][i]

    for i in index_f:
        fw_fr_1_f = D1[:,0][i]
        sp_1_f = D1[:,1][i]
        psp_sp_1_f = D1[:,2][i]
        en_fr_1_f = D1[:,3][i]
        fw_fr_2_f = D2[:,0][i]
        sp_2_f = D2[:,1][i]
        psp_sp_2_f = D2[:,2][i]
        en_fr_2_f = D2[:,3][i]
        fw_fr_3_f = D3[:,0][i]
        sp_3_f = D3[:,1][i]
        psp_sp_3_f = D3[:,2][i]
        en_fr_3_f = D3[:,3][i]
        fw_fr_4_f = D4[:,0][i]
        sp_4_f = D4[:,1][i]
        psp_sp_4_f = D4[:,2][i]
        en_fr_4_f = D4[:,3][i]

    #--------------
    #Age Seperation

    ages = D0[:,5].astype(int)            
    
    index_range_1 = np.where(ages < 26)
    index_range_2 = np.where((ages < 30) & (ages >= 26))
    index_range_3 = np.where(ages >= 30)


    for i in index_range_1:
        fw_fr_1_r1 = D1[:,0][i]
        sp_1_r1 = D1[:,1][i]
        psp_sp_1_r1 = D1[:,2][i]
        en_fr_1_r1 = D1[:,3][i]
        fw_fr_2_r1 = D2[:,0][i]
        sp_2_r1 = D2[:,1][i]
        psp_sp_2_r1 = D2[:,2][i]
        en_fr_2_r1 = D2[:,3][i]
        fw_fr_3_r1 = D3[:,0][i]
        sp_3_r1 = D3[:,1][i]
        psp_sp_3_r1 = D3[:,2][i]
        en_fr_3_r1 = D3[:,3][i]
        fw_fr_4_r1 = D4[:,0][i]
        sp_4_r1 = D4[:,1][i]
        psp_sp_4_r1 = D4[:,2][i]
        en_fr_4_r1 = D4[:,3][i]

    for i in index_range_2:
        fw_fr_1_r2 = D1[:,0][i]
        sp_1_r2 = D1[:,1][i]
        psp_sp_1_r2 = D1[:,2][i]
        en_fr_1_r2 = D1[:,3][i]
        fw_fr_2_r2 = D2[:,0][i]
        sp_2_r2 = D2[:,1][i]
        psp_sp_2_r2 = D2[:,2][i]
        en_fr_2_r2 = D2[:,3][i]
        fw_fr_3_r2 = D3[:,0][i]
        sp_3_r2 = D3[:,1][i]
        psp_sp_3_r2 = D3[:,2][i]
        en_fr_3_r2 = D3[:,3][i]
        fw_fr_4_r2 = D4[:,0][i]
        sp_4_r2 = D4[:,1][i]
        psp_sp_4_r2 = D4[:,2][i]
        en_fr_4_r2 = D4[:,3][i]

    for i in index_range_3:
        fw_fr_1_r3 = D1[:,0][i]
        sp_1_r3 = D1[:,1][i]
        psp_sp_1_r3 = D1[:,2][i]
        en_fr_1_r3 = D1[:,3][i]
        fw_fr_2_r3 = D2[:,0][i]
        sp_2_r3 = D2[:,1][i]
        psp_sp_2_r3 = D2[:,2][i]
        en_fr_2_r3 = D2[:,3][i]
        fw_fr_3_r3 = D3[:,0][i]
        sp_3_r3 = D3[:,1][i]
        psp_sp_3_r3 = D3[:,2][i]
        en_fr_3_r3 = D3[:,3][i]
        fw_fr_4_r3 = D4[:,0][i]
        sp_4_r3 = D4[:,1][i]
        psp_sp_4_r3 = D4[:,2][i]
        en_fr_4_r3 = D4[:,3][i]

    #--------------
    df1_m =  pd.DataFrame(np.array([fw_fr_1_m, sp_1_m, psp_sp_1_m, en_fr_1_m]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #83*4
    df2_m =  pd.DataFrame(np.array([fw_fr_2_m, sp_2_m, psp_sp_2_m, en_fr_2_m]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #83*4
    df3_m =  pd.DataFrame(np.array([fw_fr_3_m, sp_3_m, psp_sp_3_m, en_fr_3_m]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #83*4
    df4_m =  pd.DataFrame(np.array([fw_fr_4_m, sp_4_m, psp_sp_4_m, en_fr_4_m]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #83*4

    df1_f =  pd.DataFrame(np.array([fw_fr_1_f, sp_1_f, psp_sp_1_f, en_fr_1_f]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4
    df2_f =  pd.DataFrame(np.array([fw_fr_2_f, sp_2_f, psp_sp_2_f, en_fr_2_f]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4
    df3_f =  pd.DataFrame(np.array([fw_fr_3_f, sp_3_f, psp_sp_3_f, en_fr_3_f]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4
    df4_f =  pd.DataFrame(np.array([fw_fr_4_f, sp_4_f, psp_sp_4_f, en_fr_4_f]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4

        
    df1_r1 =  pd.DataFrame(np.array([fw_fr_1_r1, sp_1_r1, psp_sp_1_r1, en_fr_1_r1]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #32*4
    df2_r1 =  pd.DataFrame(np.array([fw_fr_2_r1, sp_2_r1, psp_sp_2_r1, en_fr_2_r1]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #32*4
    df3_r1 =  pd.DataFrame(np.array([fw_fr_3_r1, sp_3_r1, psp_sp_3_r1, en_fr_3_r1]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #32*4
    df4_r1 =  pd.DataFrame(np.array([fw_fr_4_r1, sp_4_r1, psp_sp_4_r1, en_fr_4_r1]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #32*4

    df1_r2 =  pd.DataFrame(np.array([fw_fr_1_r2, sp_1_r2, psp_sp_1_r2, en_fr_1_r2]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4
    df2_r2 =  pd.DataFrame(np.array([fw_fr_2_r2, sp_2_r2, psp_sp_2_r2, en_fr_2_r2]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4
    df3_r2 =  pd.DataFrame(np.array([fw_fr_3_r2, sp_3_r2, psp_sp_3_r2, en_fr_3_r2]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4
    df4_r2 =  pd.DataFrame(np.array([fw_fr_4_r2, sp_4_r2, psp_sp_4_r2, en_fr_4_r2]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #61*4

    df1_r3 =  pd.DataFrame(np.array([fw_fr_1_r3, sp_1_r3, psp_sp_1_r3, en_fr_1_r3]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #51*4
    df2_r3 =  pd.DataFrame(np.array([fw_fr_2_r3, sp_2_r3, psp_sp_2_r3, en_fr_2_r3]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #51*4
    df3_r3 =  pd.DataFrame(np.array([fw_fr_3_r3, sp_3_r3, psp_sp_3_r3, en_fr_3_r3]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #51*4
    df4_r3 =  pd.DataFrame(np.array([fw_fr_4_r3, sp_4_r3, psp_sp_4_r3, en_fr_4_r3]).T, columns=['fw/fr', 'sp', 'psp/sp', 'en/fr']) #51*4


    longitudinal_correlation (df1, df2, df3, df4, 'all_Correlation')

    longitudinal_correlation (df1_m, df2_m, df3_m, df4_m, 'men_Correlation')

    longitudinal_correlation (df1_f, df2_f, df3_f, df4_f, 'women_Correlation')

    longitudinal_correlation (df1_r1, df2_r1, df3_r1, df4_r1, 'Age Range 1')

    longitudinal_correlation (df1_r2, df2_r2, df3_r2, df4_r2, 'Age Range 2')

    longitudinal_correlation (df1_r3, df2_r3, df3_r3, df4_r3, 'Age Range 3')

#---------------------------------------------------    
def longitudinal_correlation (df1, df2, df3, df4, name_outfile):
    import seaborn as sns
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    from scipy.stats import spearmanr
    from scipy.stats import pearsonr

    df21 = df2-df1
    df32 = df3-df2
    df43 = df4-df3


    f = open(name_outfile + '.txt','a')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')

    #---------------------------------
    f.write('\n' + '------------------------------------------')
    f.write('\n' + 'Pearson')
    f.write('\n' + '------------------------------------------')
    f.write('\n' + '21 - 32')
    f.write('\n' + '----------------')

    coeffmat = np.zeros((df21.shape[1], df32.shape[1]))
    pvalmat = np.zeros((df21.shape[1], df32.shape[1]))
    for i in range(df21.shape[1]):    
        for j in range(df32.shape[1]):        
            corrtest = pearsonr(df21[df21.columns[i]], df32[df32.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df32.columns, index=df21.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df32.columns, index=df21.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' +  str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' +  str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '32 - 43')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df32.shape[1], df43.shape[1]))
    pvalmat = np.zeros((df32.shape[1], df43.shape[1]))
    for i in range(df32.shape[1]):    
        for j in range(df43.shape[1]):        
            corrtest = pearsonr(df32[df32.columns[i]], df43[df43.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df43.columns, index=df32.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df43.columns, index=df32.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '1 - 21')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df1.shape[1], df21.shape[1]))
    pvalmat = np.zeros((df1.shape[1], df21.shape[1]))
    for i in range(df1.shape[1]):    
        for j in range(df21.shape[1]):        
            corrtest = pearsonr(df1[df1.columns[i]], df21[df21.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df21.columns, index=df1.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df21.columns, index=df1.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '2 - 32')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df2.shape[1], df32.shape[1]))
    pvalmat = np.zeros((df2.shape[1], df32.shape[1]))
    for i in range(df2.shape[1]):    
        for j in range(df32.shape[1]):        
            corrtest = pearsonr(df2[df2.columns[i]], df32[df32.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df32.columns, index=df2.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df32.columns, index=df2.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '3 - 43')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df3.shape[1], df43.shape[1]))
    pvalmat = np.zeros((df3.shape[1], df43.shape[1]))
    for i in range(df3.shape[1]):    
        for j in range(df43.shape[1]):        
            corrtest = pearsonr(df3[df3.columns[i]], df43[df43.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df43.columns, index=df3.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df43.columns, index=df3.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------
    #---------------------------------
    f.write('\n' + '------------------------------------------')
    f.write('\n' + 'Spearman')
    f.write('\n' + '------------------------------------------')
    f.write('\n' + '21 - 32')
    f.write('\n' + '----------------')

    coeffmat = np.zeros((df21.shape[1], df32.shape[1]))
    pvalmat = np.zeros((df21.shape[1], df32.shape[1]))
    for i in range(df21.shape[1]):    
        for j in range(df32.shape[1]):        
            corrtest = spearmanr(df21[df21.columns[i]], df32[df32.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df32.columns, index=df21.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df32.columns, index=df21.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' +  str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' +  str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '32 - 43')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df32.shape[1], df43.shape[1]))
    pvalmat = np.zeros((df32.shape[1], df43.shape[1]))
    for i in range(df32.shape[1]):    
        for j in range(df43.shape[1]):        
            corrtest = spearmanr(df32[df32.columns[i]], df43[df43.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df43.columns, index=df32.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df43.columns, index=df32.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '1 - 21')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df1.shape[1], df21.shape[1]))
    pvalmat = np.zeros((df1.shape[1], df21.shape[1]))
    for i in range(df1.shape[1]):    
        for j in range(df21.shape[1]):        
            corrtest = spearmanr(df1[df1.columns[i]], df21[df21.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df21.columns, index=df1.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df21.columns, index=df1.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '2 - 32')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df2.shape[1], df32.shape[1]))
    pvalmat = np.zeros((df2.shape[1], df32.shape[1]))
    for i in range(df2.shape[1]):    
        for j in range(df32.shape[1]):        
            corrtest = spearmanr(df2[df2.columns[i]], df32[df32.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df32.columns, index=df2.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df32.columns, index=df2.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    #---------------------------------
    f.write('\n' + '3 - 43')
    f.write('\n' + '----------------')
    
    coeffmat = np.zeros((df3.shape[1], df43.shape[1]))
    pvalmat = np.zeros((df3.shape[1], df43.shape[1]))
    for i in range(df3.shape[1]):    
        for j in range(df43.shape[1]):        
            corrtest = spearmanr(df3[df3.columns[i]], df43[df43.columns[j]])  
            coeffmat[i,j] = corrtest[0]
            pvalmat[i,j] = corrtest[1]
    dfcoeff = pd.DataFrame(coeffmat, columns=df43.columns, index=df3.columns)
    dfpvals = pd.DataFrame(pvalmat, columns=df43.columns, index=df3.columns)

    f.write('\n' + 'Coefficients')
    f.write('\n' + str(dfcoeff))
    f.write('\n' + '----------------')
    f.write('\n' + 'P-Values')
    f.write('\n' + str(dfpvals))
    f.write('\n' + '----------------')
    f.write('\n' + '----------------')
    f.write('\n' + '----------------------------------------')
    #---------------------------------

    f.close()
#----------------    


#---------------------------------------------------    
main()
input("\n press enter key to exit.")
